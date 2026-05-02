#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║  URL LIVENESS CHECKER v2.0 — Flask Web GUI               ║
║  Custom ports + Screenshots + Live Console                ║
║  Usage: python3 url_checker_gui.py                        ║
║  Then open http://127.0.0.1:5001                          ║
╚══════════════════════════════════════════════════════════╝
"""
import os, sys, time, re, json, queue, threading, uuid
from pathlib import Path
from datetime import datetime
import concurrent.futures

try:
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    import socket

from flask import Flask, request, jsonify, Response, send_file
app = Flask(__name__)

UPLOAD = Path("uploads"); UPLOAD.mkdir(exist_ok=True)
REPORTS = Path("reports"); REPORTS.mkdir(exist_ok=True)
JOBS = {}

# ── Check functions ──────────────────────────────────────
def check_http(host, port, scheme, timeout=5):
    url = f"{scheme}://{host}" if port in (80,443) else f"{scheme}://{host}:{port}"
    try:
        r = requests.head(url, timeout=timeout, verify=False, allow_redirects=True,
                         headers={"User-Agent":"Mozilla/5.0 URLChecker/2.0"})
        return url, "ACTIVE", r.status_code, dict(r.headers)
    except requests.exceptions.ConnectTimeout: return url,"TIMEOUT",0,{}
    except requests.exceptions.SSLError: return url,"SSL_ERROR",0,{}
    except requests.exceptions.ConnectionError: return url,"INACTIVE",0,{}
    except Exception: return url,"ERROR",0,{}

def check_socket(host, port, scheme, timeout=3):
    url = f"{scheme}://{host}" if port in (80,443) else f"{scheme}://{host}:{port}"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(timeout); s.connect((host, port)); s.close()
        return url, "ACTIVE", 0, {}
    except socket.timeout: return url,"TIMEOUT",0,{}
    except (ConnectionRefusedError, OSError): return url,"INACTIVE",0,{}
    except Exception: return url,"ERROR",0,{}

check_fn = check_http if HAS_REQUESTS else check_socket

def load_urls(filepath):
    fp = Path(filepath)
    if fp.suffix.lower() in (".xlsx",".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(str(fp), data_only=True, read_only=True)
        ws = wb.active
        urls = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = str(row[0] or "").strip()
            if val and val.lower() not in ("url","hostname","host","domain") and "." in val:
                val = re.sub(r'^https?://','',val).split('/')[0].strip()
                if val: urls.append(val)
        wb.close()
        return list(dict.fromkeys(urls))
    else:
        urls = []
        with open(filepath) as f:
            for line in f:
                val = line.strip()
                if val and not val.startswith("#") and "." in val:
                    val = re.sub(r'^https?://','',val).split('/')[0].strip()
                    if val: urls.append(val)
        return list(dict.fromkeys(urls))

# ── Screenshot tools (cross-platform) ────────────────────
def _find_screenshot_tool():
    import shutil
    try:
        from playwright.sync_api import sync_playwright; return {"name":"playwright","type":"python"}
    except ImportError: pass
    try:
        from selenium import webdriver; return {"name":"selenium","type":"python"}
    except ImportError: pass
    for name,cmd in [("gowitness","gowitness"),("cutycapt","cutycapt"),("wkhtmltoimage","wkhtmltoimage")]:
        if shutil.which(cmd): return {"name":name,"type":"cli","cmd":cmd}
    for b in ["google-chrome","chromium-browser","chromium","chrome"]:
        if shutil.which(b): return {"name":"chrome-headless","type":"cli","cmd":b}
    for ep in [r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
               r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"]:
        if Path(ep).exists(): return {"name":"edge-headless","type":"cli","cmd":ep}
    return None

def _take_screenshot(tool, url, output_path, timeout=15):
    import subprocess as _sp
    if tool["name"]=="playwright":
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                br=p.chromium.launch(headless=True,args=["--no-sandbox","--disable-gpu"])
                pg=br.new_page(viewport={"width":1280,"height":900})
                pg.set_default_timeout(timeout*1000)
                try: pg.goto(url,wait_until="domcontentloaded",timeout=timeout*1000); pg.wait_for_timeout(1500)
                except: pass
                pg.screenshot(path=output_path,full_page=False); br.close()
            return True
        except: return False
    elif tool["name"]=="selenium":
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            opts=Options(); opts.add_argument("--headless"); opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-gpu"); opts.add_argument("--window-size=1280,900")
            opts.add_argument("--ignore-certificate-errors")
            d=webdriver.Chrome(options=opts); d.set_page_load_timeout(timeout)
            try: d.get(url); time.sleep(2)
            except: pass
            d.save_screenshot(output_path); d.quit(); return True
        except: return False
    elif tool["name"]=="gowitness":
        r=_sp.run(["gowitness","single",url,"--screenshot-path",str(Path(output_path).parent),
                   "--timeout",str(timeout)], capture_output=True, timeout=timeout+10)
        return r.returncode==0
    elif tool["name"]=="cutycapt":
        r=_sp.run(["cutycapt",f"--url={url}",f"--out={output_path}","--insecure",
                   f"--max-wait={timeout*1000}"], capture_output=True, timeout=timeout+10)
        return r.returncode==0
    elif tool["name"]=="wkhtmltoimage":
        r=_sp.run(["wkhtmltoimage","--quiet","--javascript-delay","2000","--width","1280",
                   url,output_path], capture_output=True, timeout=timeout+10)
        return r.returncode==0
    elif tool["name"] in ("chrome-headless","edge-headless"):
        r=_sp.run([tool["cmd"],"--headless","--disable-gpu","--no-sandbox",
                   "--ignore-certificate-errors",f"--screenshot={output_path}",
                   "--window-size=1280,900",url], capture_output=True, timeout=timeout+10)
        return r.returncode==0
    return False

# ── Port list builder ────────────────────────────────────
def build_port_list(ports_input):
    port_list = []; seen = set()
    for p in ports_input:
        p = int(p)
        if p in (443,8443,4443):
            key=(p,"https")
            if key not in seen: port_list.append(key); seen.add(key)
        elif p in (80,8080,3000,5000,9090):
            key=(p,"http")
            if key not in seen: port_list.append(key); seen.add(key)
        else:
            for scheme in ["http","https"]:
                key=(p,scheme)
                if key not in seen: port_list.append(key); seen.add(key)
    return port_list if port_list else [(80,"http"),(443,"https"),(8080,"http"),(8080,"https")]

# ── Scan job ─────────────────────────────────────────────
def run_scan(job_id, urls, workers=30, timeout=5, custom_ports=None, take_screenshots=False):
    q = JOBS[job_id]["queue"]
    PORTS = build_port_list(custom_ports) if custom_ports else [(80,"http"),(443,"https"),(8080,"http"),(8080,"https")]
    total = len(urls)*len(PORTS)
    results = {}; active_urls = []

    def emit(ev, d):
        q.put(json.dumps({"event":ev,"data":d}, default=str))

    emit("log",{"msg":f"Starting: {len(urls)} hosts × {len(PORTS)} ports = {total} checks","level":"info"})
    emit("log",{"msg":f"Ports: {', '.join(f'{s}({p})' for p,s in PORTS)} | Threads: {workers} | Screenshots: {'ON' if take_screenshots else 'OFF'}","level":"info"})
    emit("progress",{"done":0,"total":total,"pct":0})
    emit("ports",{"ports":[{"port":p,"scheme":s} for p,s in PORTS]})

    start = time.time(); done_count = 0; lock = threading.Lock()
    for host in urls: results[host] = {}
    tasks = [(h,p,s) for h in urls for p,s in PORTS]

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(check_fn,h,p,s,timeout):(h,p,s) for h,p,s in tasks}
        for future in concurrent.futures.as_completed(futures):
            h,p,s = futures[future]
            url,status,code,hdrs = future.result()
            results[h][f"{s}_{p}"] = {"status":status,"code":code,"url":url}
            if status == "ACTIVE": active_urls.append({"host":h,"url":url,"port":p,"scheme":s})
            with lock: done_count += 1
            icon = {"ACTIVE":"✅","INACTIVE":"❌","TIMEOUT":"⏱","SSL_ERROR":"🔒","ERROR":"💥"}.get(status,"?")
            code_str = f" ({code})" if code else ""
            level = "good" if status=="ACTIVE" else "warn" if status in ("TIMEOUT","SSL_ERROR") else "dim"
            emit("check",{"host":h,"url":url,"status":status,"code":code,"icon":icon,"port":p,"scheme":s})
            emit("log",{"msg":f"{icon} {url} → {status}{code_str}","level":level})
            if done_count % 50 == 0 or done_count == total:
                emit("progress",{"done":done_count,"total":total,"pct":round(done_count*100/total)})

    elapsed_check = time.time() - start
    emit("log",{"msg":f"Port check done in {elapsed_check:.1f}s — {len(active_urls)} active","level":"good"})

    # ── Screenshots ──────────────────────────────────────
    screenshot_count = 0
    if take_screenshots and active_urls:
        screenshot_dir = REPORTS / f"screenshots_{job_id[:8]}"
        screenshot_dir.mkdir(parents=True, exist_ok=True)
        JOBS[job_id]["screenshot_dir"] = str(screenshot_dir)
        seen = set(); targets = []
        for au in sorted(active_urls, key=lambda x:(x["host"],-x["port"],x["scheme"])):
            key = f"{au['host']}:{au['port']}"
            if key not in seen: seen.add(key); targets.append(au)
        emit("log",{"msg":f"📸 Taking {len(targets)} screenshots ({workers} threads)…","level":"info"})
        tool = _find_screenshot_tool()
        if tool:
            emit("log",{"msg":f"Tool: {tool['name']}","level":"info"})
            ss_lock = threading.Lock()
            def _ss_task(idx, t):
                nonlocal screenshot_count
                try:
                    fn = re.sub(r'[^a-zA-Z0-9._\-]','_',t["url"])[:80]+".png"
                    op = screenshot_dir / fn
                    ok = _take_screenshot(tool, t["url"], str(op), timeout=15)
                    if ok and op.exists():
                        with ss_lock: screenshot_count += 1
                        emit("screenshot",{"host":t["host"],"url":t["url"],"filename":fn})
                        emit("log",{"msg":f"  📸 {t['url']}","level":"good"})
                    return idx
                except Exception as e:
                    emit("log",{"msg":f"  Screenshot error {t['url']}: {e}","level":"warn"})
                    return idx
            # Use fewer threads for screenshots (browser instances are heavy)
            ss_workers = min(workers, 5)
            ss_done = 0
            with concurrent.futures.ThreadPoolExecutor(max_workers=ss_workers) as ss_pool:
                futures = {ss_pool.submit(_ss_task, i, t): i for i, t in enumerate(targets)}
                for future in concurrent.futures.as_completed(futures):
                    ss_done += 1
                    if ss_done % 10 == 0 or ss_done == len(targets):
                        emit("progress",{"done":done_count,"total":total,"pct":100,
                             "sub":f"Screenshots: {ss_done}/{len(targets)}"})
        else:
            emit("log",{"msg":"⚠ No screenshot tool found. Install one:","level":"warn"})
            emit("log",{"msg":"  pip install playwright && playwright install chromium","level":"info"})
            emit("log",{"msg":"  pip install selenium (+ chromedriver)","level":"info"})
            emit("log",{"msg":"  sudo apt install cutycapt | wkhtmltopdf","level":"info"})
            emit("log",{"msg":"  go install github.com/sensepost/gowitness@latest","level":"info"})

    elapsed = time.time() - start
    rows = []; live_count = 0; down_count = 0
    for host in urls:
        active = sum(1 for v in results[host].values() if v.get("status")=="ACTIVE")
        status = "LIVE" if active>0 else "DOWN"
        if active>0: live_count+=1
        else: down_count+=1
        rows.append({"host":host,"ports":results[host],"active_count":active,"status":status})

    try:
        xlsx_path = save_excel(rows, PORTS, elapsed, job_id, screenshot_count)
        JOBS[job_id]["xlsx_path"] = str(xlsx_path)
        emit("log",{"msg":f"📄 Excel: {xlsx_path.name}","level":"good"})
    except Exception as e:
        emit("log",{"msg":f"Excel error: {e}","level":"error"})

    emit("log",{"msg":f"\n{'═'*50}","level":"info"})
    emit("log",{"msg":f"  COMPLETE in {elapsed:.1f}s","level":"info"})
    emit("log",{"msg":f"  ✅ {live_count} LIVE | ❌ {down_count} DOWN | 📸 {screenshot_count} screenshots","level":"good"})
    emit("log",{"msg":f"{'═'*50}","level":"info"})

    emit("done",{"rows":rows,"live":live_count,"down":down_count,"total":len(urls),
                 "elapsed":round(elapsed,1),"screenshots":screenshot_count,
                 "xlsx":JOBS[job_id].get("xlsx_path",""),
                 "ports":[{"port":p,"scheme":s} for p,s in PORTS]})
    q.put(None)

def save_excel(rows, PORTS, elapsed, job_id, screenshot_count=0):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "URL Status"
    hf=Font(name="Arial",bold=True,color="FFFFFF",size=11)
    hfl=PatternFill("solid",fgColor="1a1a2e")
    bdr=Border(bottom=Side("thin",color="DDDDDD"),right=Side("thin",color="DDDDDD"))
    port_hdrs = [f"{s}://{'' if p in (80,443) else ':'+str(p)} ({p})" for p,s in PORTS]
    hdrs = ["#","Hostname"]+port_hdrs+["Active","Status"]
    for c,h in enumerate(hdrs,1):
        cell=ws.cell(1,c,h);cell.font=hf;cell.fill=hfl;cell.alignment=Alignment(horizontal="center");cell.border=bdr
    ws.column_dimensions['A'].width=5;ws.column_dimensions['B'].width=55
    for i,row in enumerate(rows,1):
        r=i+1;ws.cell(r,1,i).font=Font(name="Arial",color="999999",size=10)
        ws.cell(r,2,row["host"]).font=Font(name="Arial",color="1565C0",size=10)
        ci=3
        for port,scheme in PORTS:
            info=row["ports"].get(f"{scheme}_{port}",{"status":"ERROR","code":0})
            cell=ws.cell(r,ci);st=info["status"]
            if st=="ACTIVE":
                cell.value=f"ACTIVE ({info['code']})" if info["code"] else "ACTIVE"
                cell.font=Font(name="Arial",color="00BB66",bold=True,size=10)
                cell.fill=PatternFill("solid",fgColor="E8F5E9")
            elif st in ("TIMEOUT","SSL_ERROR"):
                cell.value=st.replace("_"," ");cell.font=Font(name="Arial",color="CC8800",size=10)
                cell.fill=PatternFill("solid",fgColor="FFF8E1")
            else:
                cell.value="INACTIVE";cell.font=Font(name="Arial",color="CC3333",size=10)
                cell.fill=PatternFill("solid",fgColor="FFEBEE")
            cell.alignment=Alignment(horizontal="center");cell.border=bdr;ci+=1
        ws.cell(r,ci,row["active_count"]).font=Font(name="Arial",color="00BB66" if row["active_count"] else "CC3333",bold=True,size=11)
        ws.cell(r,ci).alignment=Alignment(horizontal="center")
        sc=ws.cell(r,ci+1)
        if row["status"]=="LIVE":sc.value="LIVE";sc.font=Font(name="Arial",color="00BB66",bold=True)
        else:sc.value="DOWN";sc.font=Font(name="Arial",color="CC3333",bold=True)
        sc.alignment=Alignment(horizontal="center")
        for c in range(1,ci+2):ws.cell(r,c).border=bdr
    sr=len(rows)+3
    live=sum(1 for r in rows if r["status"]=="LIVE")
    ws.cell(sr,2,f"Total:{len(rows)} | Live:{live} | Down:{len(rows)-live} | Time:{elapsed:.1f}s | Screenshots:{screenshot_count}").font=Font(name="Arial",size=10)
    ws.freeze_panes="A2"
    out = REPORTS / f"url_check_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(out)); return out

# ── Flask Routes ─────────────────────────────────────────
@app.route("/")
def index(): return HTML_PAGE

@app.route("/api/scan", methods=["POST"])
def start_scan():
    job_id = uuid.uuid4().hex[:12]
    workers = int(request.form.get("workers",30))
    timeout = int(request.form.get("timeout",5))
    ports_str = request.form.get("ports","80,443,8080").strip()
    custom_ports = [int(p.strip()) for p in ports_str.split(",") if p.strip().isdigit()] or None
    take_screenshots = request.form.get("screenshots")=="1"
    urls = []
    f = request.files.get("file")
    if f and f.filename:
        fp = UPLOAD / f"{job_id}_{f.filename}"; f.save(str(fp)); urls = load_urls(str(fp))
    text_urls = request.form.get("urls","").strip()
    if text_urls:
        for line in text_urls.split("\n"):
            val = line.strip()
            if val and "." in val:
                val = re.sub(r'^https?://','',val).split('/')[0].strip()
                if val: urls.append(val)
    urls = list(dict.fromkeys(urls))
    if not urls: return jsonify({"error":"No URLs provided"}),400
    JOBS[job_id] = {"queue":queue.Queue(),"status":"running"}
    threading.Thread(target=run_scan,args=(job_id,urls,workers,timeout,custom_ports,take_screenshots),daemon=True).start()
    return jsonify({"job_id":job_id,"count":len(urls)})

@app.route("/api/stream/<job_id>")
def stream(job_id):
    if job_id not in JOBS: return jsonify({"error":"not found"}),404
    def generate():
        q = JOBS[job_id]["queue"]
        while True:
            try:
                msg = q.get(timeout=120)
                if msg is None: yield f"data: {json.dumps({'event':'end'})}\n\n"; break
                yield f"data: {msg}\n\n"
            except: yield f"data: {json.dumps({'event':'keepalive'})}\n\n"
    return Response(generate(),mimetype="text/event-stream",headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/api/download/<job_id>")
def download(job_id):
    if job_id not in JOBS: return jsonify({"error":"not found"}),404
    xlsx = JOBS[job_id].get("xlsx_path")
    if xlsx and Path(xlsx).exists(): return send_file(xlsx,as_attachment=True,download_name=Path(xlsx).name)
    return jsonify({"error":"not ready"}),404

@app.route("/api/screenshot/<job_id>/<filename>")
def serve_screenshot(job_id, filename):
    if job_id not in JOBS: return jsonify({"error":"not found"}),404
    sd = JOBS[job_id].get("screenshot_dir")
    if not sd: return jsonify({"error":"no screenshots"}),404
    fp = Path(sd)/filename
    if fp.exists(): return send_file(str(fp),mimetype="image/png")
    return jsonify({"error":"file not found"}),404

@app.route("/api/screenshots/<job_id>")
def gallery(job_id):
    if job_id not in JOBS: return jsonify({"error":"not found"}),404
    sd = JOBS[job_id].get("screenshot_dir")
    if not sd or not Path(sd).exists(): return "No screenshots",404
    files = sorted(Path(sd).glob("*.png"))
    html = f"""<!DOCTYPE html><html><head><title>Screenshots</title>
    <link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
    <style>body{{background:#0a0a0f;color:#e0e0e8;font-family:'JetBrains Mono',monospace;padding:20px}}
    h1{{color:#00e87b;font-size:20px;margin-bottom:16px}}
    .info{{font-size:11px;color:#6a6a7a;margin-bottom:16px}}
    .grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(420px,1fr));gap:14px}}
    .card{{background:#12121a;border:1px solid #2a2a3a;border-radius:10px;overflow:hidden;transition:transform .2s}}
    .card:hover{{transform:translateY(-2px);border-color:#00c8ff}}
    .card img{{width:100%;height:auto;display:block}}
    .card .lbl{{padding:10px 12px;font-size:10px;color:#6a6a7a;word-break:break-all;
      border-top:1px solid #2a2a3a;display:flex;justify-content:space-between}}
    .card .lbl .st{{color:#00e87b;font-weight:600}}</style></head>
    <body><h1>📸 Screenshot Gallery</h1>
    <div class="info">{len(files)} screenshots | Job: {job_id[:8]}</div><div class="grid">"""
    for f in files:
        name = f.stem.replace('_',' ')[:70]
        html += f'<div class="card"><a href="/api/screenshot/{job_id}/{f.name}" target="_blank">'
        html += f'<img src="/api/screenshot/{job_id}/{f.name}" loading="lazy"></a>'
        html += f'<div class="lbl"><span>{name}</span><span class="st">ACTIVE</span></div></div>'
    html += '</div></body></html>'; return html

# ── HTML ─────────────────────────────────────────────────
HTML_PAGE = r"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>URL Liveness Checker v2.0</title>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=DM+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{margin:0;padding:0;box-sizing:border-box}
:root{--bg:#0a0a0f;--bg2:#12121a;--bg3:#1a1a28;--bg4:#22222e;--text:#e0e0e8;--dim:#6a6a7a;--border:#2a2a3a;--green:#00e87b;--red:#ff4455;--orange:#ffaa22;--blue:#00c8ff;--purple:#b480ff;--surface:#16161f;--font:'JetBrains Mono',monospace;--sans:'DM Sans',sans-serif}
body{background:var(--bg);color:var(--text);font-family:var(--sans);min-height:100vh}
.app{max-width:1200px;margin:0 auto;padding:20px}
.header{text-align:center;padding:30px 0 20px}
.header h1{font-family:var(--font);font-size:24px;font-weight:700;letter-spacing:2px;background:linear-gradient(135deg,var(--green),var(--blue));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.header .sub{font-size:12px;color:var(--dim);letter-spacing:1px;margin-top:4px}
.sec{background:var(--bg2);border:1px solid var(--border);border-radius:12px;padding:20px;margin-bottom:16px}
.sec h2{font-size:14px;color:var(--blue);margin-bottom:12px;font-family:var(--font)}
.row{display:flex;gap:12px;flex-wrap:wrap}.box{flex:1;min-width:250px}
.box label{font-size:10px;color:var(--dim);text-transform:uppercase;letter-spacing:1px;display:block;margin-bottom:5px}
.box textarea{width:100%;height:110px;background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:10px;color:var(--text);font-family:var(--font);font-size:11px;resize:vertical}
.box textarea:focus{outline:none;border-color:var(--blue)}
.box input[type=file]{width:100%;padding:10px;background:var(--bg);border:2px dashed var(--border);border-radius:8px;color:var(--dim);font-size:11px;cursor:pointer}
.cfg{display:flex;gap:10px;margin-top:12px;align-items:center;flex-wrap:wrap}
.cfg label{font-size:10px;color:var(--dim)}
.cfg input[type=number]{width:55px;padding:4px 6px;background:var(--bg);border:1px solid var(--border);border-radius:6px;color:var(--text);font-family:var(--font);font-size:11px;text-align:center}
.cfg input[type=text]{width:220px;padding:4px 8px;background:var(--bg);border:1px solid var(--border);border-radius:6px;color:var(--text);font-family:var(--font);font-size:11px}
.pre{font-size:9px;cursor:pointer;color:var(--blue);text-decoration:underline}.pre:hover{color:var(--green)}
.chk{display:flex;align-items:center;gap:4px}.chk input{accent-color:var(--green)}.chk label{font-size:10px;color:var(--blue);cursor:pointer}
.btn{padding:10px 28px;background:linear-gradient(135deg,var(--green),#00b86e);color:#000;border:none;border-radius:8px;font-family:var(--font);font-size:13px;font-weight:700;cursor:pointer;letter-spacing:1px;transition:all .2s;margin-left:auto}
.btn:hover{transform:translateY(-1px);box-shadow:0 4px 20px rgba(0,232,123,.3)}
.btn:disabled{opacity:.4;cursor:not-allowed;transform:none}.btn.running{background:linear-gradient(135deg,var(--red),#cc3344);color:#fff}
.stats{display:flex;gap:8px;margin-bottom:16px}
.st{flex:1;background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:14px;text-align:center}
.st .v{font-family:var(--font);font-size:28px;font-weight:700}.st .l{font-size:9px;color:var(--dim);text-transform:uppercase;letter-spacing:1.5px;margin-top:2px}
.st-live .v{color:var(--green)}.st-down .v{color:var(--red)}.st-total .v{color:var(--blue)}.st-time .v{color:var(--orange)}.st-ss .v{color:var(--purple)}
.prog{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:12px 16px;margin-bottom:16px;display:none}
.pbar{height:6px;background:var(--bg);border-radius:3px;overflow:hidden;margin-bottom:6px}
.pfill{height:100%;background:linear-gradient(90deg,var(--green),var(--blue));width:0%;transition:width .3s;border-radius:3px}
.ptxt{font-size:10px;color:var(--dim);font-family:var(--font);display:flex;justify-content:space-between}
.res{background:var(--bg2);border:1px solid var(--border);border-radius:12px;overflow:hidden;margin-bottom:16px;display:none}
.res-h{padding:14px 18px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.res-h h2{font-size:14px;color:var(--blue);font-family:var(--font)}
.fb{padding:3px 10px;font-size:10px;background:var(--bg);border:1px solid var(--border);border-radius:20px;color:var(--dim);cursor:pointer;font-family:var(--font)}
.fb.on{background:var(--blue);color:#000;border-color:var(--blue)}.fb:hover{border-color:var(--blue)}
.dl{padding:6px 14px;background:var(--green);color:#000;border:none;border-radius:6px;font-family:var(--font);font-size:11px;font-weight:600;cursor:pointer;display:none}
.dl:hover{background:#00ff88}.dl-ss{background:var(--purple)}.dl-ss:hover{background:#c8a0ff}
.tw{max-height:60vh;overflow-y:auto}
table{width:100%;border-collapse:collapse}
th{background:var(--bg3);font-size:10px;color:var(--dim);text-transform:uppercase;letter-spacing:1px;padding:8px 12px;text-align:center;font-family:var(--font);font-weight:600;border-bottom:1px solid var(--border);position:sticky;top:0}
th:nth-child(2){text-align:left}
td{padding:6px 12px;border-bottom:1px solid rgba(255,255,255,.03);font-size:11px;text-align:center;font-family:var(--font)}
td:nth-child(2){text-align:left;color:var(--blue)}
tr:hover{background:rgba(0,200,255,.03)}
.s-a{color:var(--green);font-weight:600}.s-i{color:var(--red)}.s-t{color:var(--orange)}.s-live{color:var(--green);font-weight:700}.s-down{color:var(--red);font-weight:700}
.con{background:var(--bg2);border:1px solid var(--border);border-radius:12px;overflow:hidden}
.con-h{padding:10px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:8px}
.con-h h2{font-size:13px;color:var(--purple);font-family:var(--font)}
.live{font-size:9px;color:var(--green);display:none;animation:pulse 1.5s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}
.clog{height:280px;overflow-y:auto;padding:10px 14px;background:var(--bg);font-family:var(--font);font-size:10px;line-height:1.7}
.log-good{color:var(--green)}.log-warn{color:var(--orange)}.log-error{color:var(--red)}.log-info{color:var(--dim)}.log-dim{color:#444}
.clr{margin-left:auto;padding:3px 10px;font-size:9px;background:var(--bg);border:1px solid var(--border);border-radius:4px;color:var(--dim);cursor:pointer;font-family:var(--font)}
</style></head><body><div class="app">
<div class="header"><h1>⬡ URL LIVENESS CHECKER</h1><div class="sub">CUSTOM PORTS · SCREENSHOTS · LIVE CONSOLE · v2.0</div></div>
<div class="sec"><h2>📋 Input</h2>
<div class="row"><div class="box"><label>Paste URLs (one per line)</label><textarea id="url-text" placeholder="example.com&#10;api.example.com&#10;staging.example.com"></textarea></div>
<div class="box"><label>Or upload (.txt / .xlsx)</label><input type="file" id="url-file" accept=".txt,.csv,.xlsx,.xls"><div id="file-tag" style="margin-top:6px;font-size:10px;color:var(--dim)"></div></div></div>
<div class="cfg">
<label>Ports:</label><input type="text" id="cfg-ports" value="80,443,8080,8443">
<span class="pre" onclick="document.getElementById('cfg-ports').value='80,443'">Web</span>
<span class="pre" onclick="document.getElementById('cfg-ports').value='80,443,8080,8443,3000,5000,9090'">Extended</span>
<span class="pre" onclick="document.getElementById('cfg-ports').value='21,22,25,53,80,110,143,443,445,993,995,1433,3306,3389,5432,8080,8443'">Full</span>
<label style="margin-left:8px">Threads:</label><input type="number" id="cfg-w" value="30" min="1" max="100">
<label>Timeout:</label><input type="number" id="cfg-t" value="5" min="1" max="30">
<div class="chk" style="margin-left:8px"><input type="checkbox" id="cfg-ss" checked><label for="cfg-ss">📸 Screenshots</label></div>
<button class="btn" id="scan-btn" onclick="go()">▶ START SCAN</button>
</div></div>
<div class="stats">
<div class="st st-total"><div class="v" id="v-total">—</div><div class="l">Total</div></div>
<div class="st st-live"><div class="v" id="v-live">—</div><div class="l">Live</div></div>
<div class="st st-down"><div class="v" id="v-down">—</div><div class="l">Down</div></div>
<div class="st st-time"><div class="v" id="v-time">—</div><div class="l">Seconds</div></div>
<div class="st st-ss"><div class="v" id="v-ss">—</div><div class="l">Screenshots</div></div>
</div>
<div class="prog" id="prog"><div class="pbar"><div class="pfill" id="pfill"></div></div><div class="ptxt"><span id="plbl">Starting…</span><span id="ppct">0%</span></div></div>
<div class="res" id="res"><div class="res-h"><h2>📊 Results</h2>
<div style="display:flex;gap:4px;margin-left:8px"><span class="fb on" onclick="filt('all',this)">All</span><span class="fb" onclick="filt('LIVE',this)">✅ Live</span><span class="fb" onclick="filt('DOWN',this)">❌ Down</span></div>
<div style="margin-left:auto;display:flex;gap:6px"><button class="dl" id="dl-x" onclick="dlX()">⬇ Excel</button><button class="dl dl-ss" id="dl-s" onclick="dlS()">📸 Gallery</button></div>
</div><div class="tw"><table><thead><tr id="th"></tr></thead><tbody id="tb"></tbody></table></div></div>
<div class="con"><div class="con-h"><h2>🖥 Console</h2><span class="live" id="live">● LIVE</span><button class="clr" onclick="document.getElementById('c').innerHTML=''">✕ Clear</button></div><div class="clog" id="c"><div class="log-dim">Ready.</div></div></div>
</div>
<script>
let J=null,R=[],F='all',P=[];
document.getElementById('url-file').onchange=function(){const f=this.files[0];document.getElementById('file-tag').textContent=f?'✓ '+f.name+' ('+(f.size/1024).toFixed(1)+'KB)':'';};
function L(m,l='info'){const el=document.getElementById('c');const d=document.createElement('div');d.className='log-'+l;d.textContent='['+new Date().toLocaleTimeString()+'] '+m;el.appendChild(d);el.scrollTop=el.scrollHeight;}
function go(){
  const btn=document.getElementById('scan-btn');if(btn.classList.contains('running'))return;
  const fd=new FormData();const t=document.getElementById('url-text').value.trim();
  if(t)fd.append('urls',t);const f=document.getElementById('url-file').files[0];
  if(f)fd.append('file',f);if(!t&&!f){L('Provide URLs','error');return;}
  fd.append('ports',document.getElementById('cfg-ports').value);
  fd.append('workers',document.getElementById('cfg-w').value);
  fd.append('timeout',document.getElementById('cfg-t').value);
  if(document.getElementById('cfg-ss').checked)fd.append('screenshots','1');
  btn.disabled=true;btn.classList.add('running');btn.textContent='⏹ SCANNING…';
  document.getElementById('prog').style.display='block';document.getElementById('live').style.display='inline';
  document.getElementById('res').style.display='block';document.getElementById('tb').innerHTML='';
  document.getElementById('dl-x').style.display='none';document.getElementById('dl-s').style.display='none';
  R=[];P=[];document.getElementById('c').innerHTML='';L('Uploading…');
  fetch('/api/scan',{method:'POST',body:fd}).then(r=>r.json()).then(d=>{
    if(d.error){L('Error: '+d.error,'error');rst();return;}
    J=d.job_id;L('Job '+J+': '+d.count+' hosts');
    const es=new EventSource('/api/stream/'+J);
    es.onmessage=e=>{try{const m=JSON.parse(e.data);H(m);if(m.event==='end')es.close();}catch(x){}};
    es.onerror=()=>{L('Connection lost','error');es.close();rst();};
  }).catch(e=>{L('Failed: '+e,'error');rst();});
}
function rst(){const b=document.getElementById('scan-btn');b.disabled=false;b.classList.remove('running');b.textContent='▶ START SCAN';document.getElementById('live').style.display='none';}
function H(m){const{event,data}=m;
  if(event==='log')L(data.msg,data.level);
  else if(event==='ports'){P=data.ports;document.getElementById('th').innerHTML='<th>#</th><th>Hostname</th>'+P.map(p=>'<th>'+p.scheme+'('+p.port+')</th>').join('')+'<th>Active</th><th>Status</th>';}
  else if(event==='progress'){document.getElementById('pfill').style.width=data.pct+'%';document.getElementById('plbl').textContent=data.sub||data.done+'/'+data.total;document.getElementById('ppct').textContent=data.pct+'%';}
  else if(event==='screenshot')L('📸 '+data.url,'good');
  else if(event==='done'){R=data.rows||[];P=data.ports||P;document.getElementById('v-total').textContent=data.total;document.getElementById('v-live').textContent=data.live;document.getElementById('v-down').textContent=data.down;document.getElementById('v-time').textContent=data.elapsed+'s';document.getElementById('v-ss').textContent=data.screenshots||0;document.getElementById('dl-x').style.display='inline-block';if(data.screenshots>0)document.getElementById('dl-s').style.display='inline-block';rend();rst();}
}
function rend(){const tb=document.getElementById('tb');const f=F==='all'?R:R.filter(r=>r.status===F);
  tb.innerHTML=f.map((r,i)=>{const cells=P.map(p=>{const k=p.scheme+'_'+p.port;const info=r.ports[k]||{status:'ERROR',code:0};const cls={ACTIVE:'s-a',INACTIVE:'s-i',TIMEOUT:'s-t',SSL_ERROR:'s-t'}[info.status]||'s-i';const lbl=info.status==='ACTIVE'?(info.code?'ACTIVE('+info.code+')':'ACTIVE'):info.status.replace('_',' ');return '<td class="'+cls+'">'+lbl+'</td>';}).join('');return '<tr><td style="color:var(--dim)">'+(i+1)+'</td><td>'+r.host+'</td>'+cells+'<td style="color:'+(r.active_count?'var(--green)':'var(--red)')+';font-weight:700">'+r.active_count+'</td><td class="s-'+r.status.toLowerCase()+'">'+r.status+'</td></tr>';}).join('');
}
function filt(f,btn){F=f;document.querySelectorAll('.fb').forEach(b=>b.classList.remove('on'));if(btn)btn.classList.add('on');rend();}
function dlX(){if(J)window.location.href='/api/download/'+J;}
function dlS(){if(J)window.open('/api/screenshots/'+J,'_blank');}
</script></body></html>"""

if __name__ == "__main__":
    print("╔══════════════════════════════════════════════════════╗")
    print("║  URL LIVENESS CHECKER v2.0                           ║")
    print("║  http://127.0.0.1:5001                               ║")
    print("║  Custom ports · Screenshots · Live console            ║")
    print("╚══════════════════════════════════════════════════════╝")
    app.run(host="0.0.0.0", port=5001, debug=False, threaded=True)
